#!/usr/bin/env python
"""Gera data/schedule.json a partir do arquivo-base do cronograma."""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PRIORITY_BY_RGB = {
    "FFEBFBFF": ("blue", 1, "Azul"),
    "FFE0F6D9": ("green", 2, "Verde"),
    "FFFFF5D9": ("yellow", 3, "Amarelo"),
    "FFF9CAD0": ("red", 4, "Vermelho"),
}
DEFAULT_PRIORITY = ("neutral", 5, "Sem cor")

BLOCK_RE = re.compile(r"^Bloco\s+([\d-]+)\s+[—-]\s+Liberação:\s*(.+)$", re.IGNORECASE)
DATE_RE = re.compile(r"\b(\d{2}/\d{2}/\d{4})\b")
RATIO_RE = re.compile(r"^\s*(\d+)\s*/\s*(\d+)\s*$")


def find_default_excel() -> Path:
    candidates = [
        Path(r"C:\Users\Arthur\Desktop\s\Cronograma - Residência.xlsx"),
        Path(r"C:\Users\Arthur\Desktop\s\Cronograma - Residencia.xlsx"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate

    folder = Path(r"C:\Users\Arthur\Desktop\s")
    if folder.exists():
        matches = sorted(folder.glob("Cronograma*.xlsx"))
        if matches:
            return matches[0]

    raise FileNotFoundError("Não encontrei o arquivo de cronograma (.xlsx).")


def parse_ratio(value: Any) -> dict[str, int]:
    if value is None:
        return {"correct": 0, "total": 0}

    text = str(value).strip()
    if not text:
        return {"correct": 0, "total": 0}

    match = RATIO_RE.match(text)
    if match:
        correct = int(match.group(1))
        total = int(match.group(2))
        if correct > total:
            total = correct
        return {"correct": correct, "total": total}

    try:
        number = int(float(text))
        return {"correct": 0, "total": max(number, 0)}
    except ValueError:
        return {"correct": 0, "total": 0}


def split_area_topic(raw: str) -> tuple[str, str]:
    for sep in (" - ", ":"):
        if sep in raw:
            left, right = raw.split(sep, 1)
            left = left.strip()
            right = right.strip()
            if left and right:
                return left, right
    return "Geral", raw.strip()


def parse_block_header(label: str) -> tuple[str, list[str]]:
    match = BLOCK_RE.match(label.strip())
    if not match:
        return "?", []

    block_number = match.group(1)
    date_tokens = DATE_RE.findall(match.group(2))
    dates_iso: list[str] = []

    for token in date_tokens:
        try:
            dt = datetime.strptime(token, "%d/%m/%Y").date()
            dates_iso.append(dt.isoformat())
        except ValueError:
            continue

    return block_number, dates_iso


def normalize_rgb(cell_rgb: Any) -> str:
    value = str(cell_rgb or "").strip().upper()
    if value in {"00000000", "NONE", ""}:
        return ""
    return value


def build_seed(
    seen: Any,
    q1: dict[str, int],
    q2: dict[str, int],
    q3: dict[str, int],
    q4: dict[str, int],
) -> dict[str, dict[str, Any]]:
    has_review_data = any((q2["total"], q3["total"], q4["total"]))
    study_done = bool(seen) or q1["total"] > 0 or has_review_data

    return {
        "study": {
            "done": study_done,
            "correct": q1["correct"],
            "total": q1["total"],
        },
        "rev1w": {
            "done": q2["total"] > 0,
            "correct": q2["correct"],
            "total": q2["total"],
        },
        "rev1m": {
            "done": q3["total"] > 0,
            "correct": q3["correct"],
            "total": q3["total"],
        },
        "rev3m": {
            "done": q4["total"] > 0,
            "correct": q4["correct"],
            "total": q4["total"],
        },
        "rev6m": {
            "done": False,
            "correct": 0,
            "total": 0,
        },
    }


def parse_sheet(workbook_path: Path, sheet_name: str) -> dict[str, Any]:
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    block_rows: list[tuple[int, str]] = []
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, str) and value.strip().startswith("Bloco "):
            block_rows.append((row, value.strip()))

    if not block_rows:
        raise ValueError("Nenhum bloco encontrado na planilha.")

    block_rows.append((ws.max_row + 1, "END"))

    blocks: list[dict[str, Any]] = []
    topics: list[dict[str, Any]] = []
    topic_counter = 1

    for idx in range(len(block_rows) - 1):
        row_start, label = block_rows[idx]
        row_end, _ = block_rows[idx + 1]

        block_number, release_dates = parse_block_header(label)
        block_id = f"b{block_number.replace('-', '_')}"
        week_start = release_dates[0] if release_dates else None

        subjects: list[dict[str, Any]] = []
        for row in range(row_start + 1, row_end):
            raw = ws.cell(row, 1).value
            if not isinstance(raw, str):
                continue

            raw = raw.strip()
            if not raw:
                continue
            if raw.lower() == "assunto":
                continue
            if raw.startswith("Bloco "):
                continue

            area, topic = split_area_topic(raw)

            rgb = normalize_rgb(ws.cell(row, 1).fill.fgColor.rgb)
            priority_name, priority_rank, priority_label = PRIORITY_BY_RGB.get(rgb, DEFAULT_PRIORITY)

            q1 = parse_ratio(ws.cell(row, 3).value)
            q2 = parse_ratio(ws.cell(row, 4).value)
            q3 = parse_ratio(ws.cell(row, 5).value)
            q4 = parse_ratio(ws.cell(row, 6).value)

            topic_id = f"t{topic_counter:04d}"
            topic_counter += 1

            subjects.append(
                {
                    "id": topic_id,
                    "blockId": block_id,
                    "blockNumber": block_number,
                    "blockLabel": label,
                    "weekLabel": f"Semana {block_number}",
                    "releaseDates": release_dates,
                    "weekStartDate": week_start,
                    "plannedDate": week_start,
                    "sourceRow": row,
                    "rawTitle": raw,
                    "area": area,
                    "topic": topic,
                    "priority": priority_name,
                    "priorityRank": priority_rank,
                    "priorityLabel": priority_label,
                    "priorityColorRgb": rgb,
                    "seed": build_seed(ws.cell(row, 2).value, q1, q2, q3, q4),
                }
            )

        subjects.sort(key=lambda item: (item["priorityRank"], item["sourceRow"]))

        total_subjects = len(subjects)
        for position, item in enumerate(subjects):
            item["priorityPosition"] = position + 1
            if week_start and total_subjects > 0:
                offset = (position * 7) // total_subjects
                if offset > 6:
                    offset = 6
                dt = datetime.fromisoformat(week_start).date() + timedelta(days=offset)
                item["plannedDate"] = dt.isoformat()
            topics.append(item)

        blocks.append(
            {
                "id": block_id,
                "blockNumber": block_number,
                "label": label,
                "weekLabel": f"Semana {block_number}",
                "releaseDates": release_dates,
                "weekStartDate": week_start,
                "topicIds": [item["id"] for item in subjects],
                "topicCount": len(subjects),
                "priorityCounts": {
                    "blue": sum(1 for item in subjects if item["priority"] == "blue"),
                    "green": sum(1 for item in subjects if item["priority"] == "green"),
                    "yellow": sum(1 for item in subjects if item["priority"] == "yellow"),
                    "red": sum(1 for item in subjects if item["priority"] == "red"),
                    "neutral": sum(1 for item in subjects if item["priority"] == "neutral"),
                },
            }
        )

    areas = sorted({item["area"] for item in topics})

    return {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "sourceFile": str(workbook_path),
            "sourceSheet": sheet_name,
            "priorityLegend": {
                "blue": "Azul - maior prioridade",
                "green": "Verde - segunda prioridade",
                "yellow": "Amarelo - terceira prioridade",
                "red": "Vermelho - quarta prioridade",
                "neutral": "Sem cor mapeada - prioridade residual",
            },
            "revisionOffsetsDays": {
                "rev1w": 7,
                "rev1m": 30,
                "rev3m": 90,
                "rev6m": 180,
            },
            "topicCount": len(topics),
            "blockCount": len(blocks),
            "areaCount": len(areas),
        },
        "areas": areas,
        "blocks": blocks,
        "topics": topics,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Converte o cronograma em JSON para o site.")
    parser.add_argument("--xlsx", type=str, default=None, help="Caminho do arquivo .xlsx")
    parser.add_argument("--sheet", type=str, default="Arthur", help="Nome da aba")
    parser.add_argument(
        "--out",
        type=str,
        default=None,
        help="Arquivo de saída JSON (padrão: data/schedule.json)",
    )
    args = parser.parse_args()

    workbook_path = Path(args.xlsx) if args.xlsx else find_default_excel()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {workbook_path}")

    default_out = Path(__file__).resolve().parent / "data" / "schedule.json"
    out_path = Path(args.out) if args.out else default_out
    out_path.parent.mkdir(parents=True, exist_ok=True)

    payload = parse_sheet(workbook_path, args.sheet)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"JSON gerado em: {out_path}")
    print(f"Assuntos: {payload['meta']['topicCount']} | Blocos: {payload['meta']['blockCount']}")


if __name__ == "__main__":
    main()
